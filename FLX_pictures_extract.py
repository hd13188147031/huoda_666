import requests
import random
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Font,Alignment
from datetime import datetime
import urllib3
import re
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import uuid

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
session = requests.Session()
retries = Retry(total=5, backoff_factor=0.1, status_forcelist=[500, 502, 503, 504])
session.mount('https://', HTTPAdapter(max_retries=retries))

def password():
    mac = ':'.join(['{:02x}'.format((uuid.getnode() >> elements) & 0xff) for elements in range(2, 7)][::-1])
    password_1 = mac[8:10]
    password_2 = mac.replace(password_1, '')
    # 取出奇数位上的字符
    odd_chars = password_2[1::2]  # 前
    # 取出偶数位上的字符
    even_chars = password_2[0::2]  # 后
    key = odd_chars + even_chars
    return key,mac

def TIME():
    global star,end
    star = input('输入开始时间:(时间格式为"YYYY-MM-DD")\n')
    end = input('输入结束时间:(时间格式为"YYYY-MM-DD")\n')
    star_time = datetime.strptime(star, "%Y-%m-%d")
    end_time = datetime.strptime(end, "%Y-%m-%d")
    # 开始和结束时间的时间戳
    star_time = str(star_time.timestamp())[:-2]
    end_time = str(end_time.timestamp())[:-2]
    date = {"time_from": int(star_time), "time_to": int(end_time)}
    date_star = date.get("time_from")
    date_end = date.get("time_to")
    current_timestamp = time.time()
    a = str(current_timestamp)[:10]
    b = str(current_timestamp)[11:]
    date_string = a + b
    return date_star,date_end,date_string

def hds():
    user_agent_list = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36/',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36'
    ]
    header = random.choice(user_agent_list)
    headers = {
        "User-Agent": header,
        "Connection":'close',
        "Referer":"https://search.gd.gov.cn",
        "Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "Accept-Language":"zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6"
    }
    print(f'谷歌浏览器版本索引：{user_agent_list.index(header)}')# 试用了哪个谷歌浏览器
    return headers


def url_clean():
    # date_star,date_end,_ = TIME()
    start_url = 'https://search.gd.gov.cn/api/search/all'
    url_list = []
    num = 1
    while True:
        time.sleep(0.6)
        global parmas
        parmas = {
            "gdbsDivision": 440300,
            "keywords": " ",
            "page": num,
            "position": 'title',
            "range": 'site',
            "recommand": 1,
            "service_area": 755,
            "site_id": 755038,
            "sort": "smart",
            "time_from": date_star,
            "time_to": date_end
        }
        print(f'{"-"*24}第{num}页{"-"*24}')
        num += 1
        response = requests.post(url=start_url,params=parmas,headers=hds(),timeout=10,verify=False)
        response.close()
        data = response.json()
        data_list = data['data']['news']['list']
        # print(len(data_list))
        for dic in data_list:
            title = dic["title"] # 新闻名称
            url = dic["post_url"]# 新闻链接
            print(url)
            url_list.append(url)
        if len(data_list)<20:
            print('==========翻页结束=========')
            break
    print()
    print()
    print(url_list)
    print(f'翻页结束\n'
          f'总计采集新闻连接数量：{len(url_list)}')
    return url_list

def pictures():
    list_picture_dict = []
    for url in url_clean() :
        print(url)
        pattern = '\d+'
        a = url.rfind('/')  # 连接中最后一个/的位置
        b = url.rfind('.')  # 连接中最后一个.的位置
        try:
            d = re.findall(pattern, url[a + 1:b])[0]  # 获取url中的编码数字
            url_head = url[:url.index('/', 8) + 1] # 获取协议和域名
            if parmas['site_id'] ==755038: # 确定当前访问的是否是大鹏区
                if url_head.index('s')==4: # 如果是确认协议是否是https
                    http_head = url_head.replace('https','http') # 如果是将https换成http
                    new_url = f'{http_head}postmeta/p/{d[:-6]}/{d[:-3]}/{d}.json'
                    response = session.get(url=new_url, headers=hds(), verify=False, timeout=10)
            else:
                new_url = f'{url_head}postmeta/p/{d[:-6]}/{d[:-3]}/{d}.json'
                response = session.get(url=new_url, headers=hds(), verify=False, timeout=10)
            time.sleep(0.6)

            response.close()
            print(f'浏览器响应值：{response.status_code}')
            if response.status_code == 200:
                data = response.json()['content']
                print(f'文本内容：{data}')
                obj = re.compile(r'<img class="nfw-cms-img" img-id=".*?" src="(?P<picture>.*?)" alt="', re.S)
                tp = obj.finditer(data)
                for i in tp:
                    tpp = i.group('picture').strip()
                    print(f'图片url：{tpp}')
                    picture_dict = [tpp, url, '']
                    list_picture_dict.append(picture_dict)
            else:
                print(f"{url}请求失败，状态码：{response.status_code}")
        except requests.exceptions.ConnectionError:
            print(f'访问受限：{url_head}')
        except:
            print('连接错误')
    print(list_picture_dict)
    print(f'-----------图片数量共计：{len(list_picture_dict)}条-------------')
    print('--------------采集完毕---------------')
    return list_picture_dict

def install():
    try:
        wb = Workbook()
        ws = wb.active
        headers = ['图片链接', '来源页', '备注']
        ws.append(headers)
        # 设置表头底色和字体颜色
        fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
        FONT = Font(color='FFFFFF',bold=True)
        alignment = Alignment(horizontal='center')  # 字体水平居中
        for cell in ws[1]:
            cell.fill = fill # 改变底色
            cell.font = FONT # 白色字体
            cell.alignment = alignment
        for column in 'ABC':  # 例如，设置B到F列的宽度
            ws.column_dimensions[column].width = 60
        for row in pictures():
            ws.append(row)
        # 保存Excel文件
        add = {'755038':'大鹏区','755039':'福田区','755527':'罗湖区','755042':'盐田区','755040':'南山区',
               '755037':'宝安区','755043':'龙岗区','755044':'龙华区','755041':'坪山区','755046':'光明区'}
        name = add.get(str(parmas['site_id']))
        wb.save(f"{name}图片_（{star}_{end}）.xlsx")
    except PermissionError:
        print('请关闭excle文档')

try:
    key,mac = password()
    with open('./id.txt', 'w') as f:
        f.write(mac)
    with open('./key.txt','r') as r:
        readkey = r.read()

    if readkey==key:
        date_star, date_end, date_string = TIME()
        install()
    else:
        print('密码错误请联系管理员')
except FileNotFoundError:
    print('缺少授权文件请联系管理员')

